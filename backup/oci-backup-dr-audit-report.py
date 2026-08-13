#!/usr/bin/env python3
"""Read-only OCI Backup & DR audit report for OCI Cloud Shell.

Collects accessible compartments, backup policies and schedules, policy
assignments, instances, boot/block volumes, volume groups, Volume Group
backups, and source-side replication toward a DR region. Generates CSV files
and an optional XLSX workbook in the user's HOME directory.

No OCI create/update/delete/restore API is called.
"""

import argparse
import csv
from collections import defaultdict
from datetime import datetime, timezone
from pathlib import Path

import oci

ERRORS = []


def get_all(label, func, *args, **kwargs):
    try:
        return oci.pagination.list_call_get_all_results(func, *args, **kwargs).data
    except Exception as exc:
        ERRORS.append({"operation": label, "error": str(exc)})
        print(f"[WARN] {label}: {exc}")
        return []


def make_client(cls, config, region):
    cfg = dict(config)
    cfg["region"] = region
    return cls(cfg)


def retention_days(seconds):
    return round(seconds / 86400, 2) if seconds is not None else ""


def iso(value):
    return value.isoformat() if value else ""


def write_csv(path, rows):
    if not rows:
        return
    fields = []
    for row in rows:
        for key in row:
            if key not in fields:
                fields.append(key)
    with path.open("w", newline="", encoding="utf-8-sig") as handle:
        writer = csv.DictWriter(handle, fieldnames=fields)
        writer.writeheader()
        writer.writerows(rows)


def write_xlsx(path, sheets):
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font
        from openpyxl.utils import get_column_letter
    except ImportError:
        print("[WARN] openpyxl not installed; CSV output is still available.")
        print("       Optional: python3 -m pip install --user openpyxl")
        return False

    wb = Workbook()
    wb.remove(wb.active)
    for name, rows in sheets:
        ws = wb.create_sheet(name[:31])
        if not rows:
            ws.append(["No data returned"])
            continue
        fields = []
        for row in rows:
            for key in row:
                if key not in fields:
                    fields.append(key)
        ws.append(fields)
        for cell in ws[1]:
            cell.font = Font(bold=True)
        for row in rows:
            ws.append([row.get(key, "") for key in fields])
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions
        for idx, field in enumerate(fields, 1):
            width = max([len(str(field))] + [len(str(r.get(field, ""))) for r in rows])
            ws.column_dimensions[get_column_letter(idx)].width = min(width + 2, 60)
    wb.save(path)
    return True


def parse_args():
    parser = argparse.ArgumentParser(description="OCI Backup & DR read-only audit")
    parser.add_argument("--source-region", default="sa-saopaulo-1")
    parser.add_argument("--dr-region", default="sa-vinhedo-1")
    parser.add_argument("--output-dir", default=str(Path.home()))
    parser.add_argument("--prefix", default="oci_backup_dr_audit")
    return parser.parse_args()


def main():
    args = parse_args()
    output_dir = Path(args.output_dir).expanduser().resolve()
    output_dir.mkdir(parents=True, exist_ok=True)

    # Native Cloud Shell authentication: ~/.oci/config is already available.
    config = oci.config.from_file()
    tenancy_id = config["tenancy"]
    identity = make_client(oci.identity.IdentityClient, config, args.source_region)
    tenancy = identity.get_tenancy(tenancy_id).data

    compartments = get_all(
        "list compartments",
        identity.list_compartments,
        tenancy_id,
        compartment_id_in_subtree=True,
        access_level="ACCESSIBLE",
    )
    scope = [(tenancy_id, tenancy.name)] + [
        (c.id, c.name) for c in compartments if c.lifecycle_state == "ACTIVE"
    ]

    data = defaultdict(list)
    policy_by_asset = {}
    instance_by_id = {}
    boot_to_instance = {}
    volume_to_instance = {}

    print(f"[INFO] Tenancy: {tenancy.name}")
    print(f"[INFO] Source: {args.source_region} | DR: {args.dr_region}")
    print(f"[INFO] Accessible scope: {len(scope)} compartments including root")

    block = make_client(oci.core.BlockstorageClient, config, args.source_region)
    compute = make_client(oci.core.ComputeClient, config, args.source_region)

    # Backup policies, schedules and asset assignments across the tenancy.
    for compartment_id, compartment_name in scope:
        for policy in get_all(
            f"policies/{compartment_name}",
            block.list_volume_backup_policies,
            compartment_id=compartment_id,
        ):
            data["policies"].append({
                "region": args.source_region,
                "compartment": compartment_name,
                "policy": policy.display_name,
                "policy_ocid": policy.id,
                "destination_region": getattr(policy, "destination_region", None) or "",
                "schedule_count": len(policy.schedules or []),
            })
            for number, schedule in enumerate(policy.schedules or [], 1):
                data["schedules"].append({
                    "compartment": compartment_name,
                    "policy": policy.display_name,
                    "schedule": number,
                    "backup_type": schedule.backup_type,
                    "period": schedule.period,
                    "day_of_week": schedule.day_of_week or "",
                    "day_of_month": schedule.day_of_month or "",
                    "month": schedule.month or "",
                    "hour_of_day": schedule.hour_of_day,
                    "time_zone": schedule.time_zone or "",
                    "retention_days": retention_days(schedule.retention_seconds),
                    "prevent_deletion": getattr(schedule, "is_prevent_deletion_enabled", None),
                    "retention_lock": getattr(schedule, "is_retention_lock_enabled", None),
                })
            for assignment in get_all(
                f"assignments/{policy.display_name}",
                block.get_volume_backup_policy_asset_assignment,
                policy_id=policy.id,
            ):
                asset_id = assignment.asset_id
                asset_type = asset_id.split(".")[1] if asset_id and "." in asset_id else "unknown"
                policy_by_asset[asset_id] = policy.display_name
                data["assignments"].append({
                    "compartment": compartment_name,
                    "policy": policy.display_name,
                    "asset_type": asset_type,
                    "asset_ocid": asset_id,
                })

    # Compute inventory and volume attachments.
    for compartment_id, compartment_name in scope:
        for instance in get_all(
            f"instances/{compartment_name}", compute.list_instances, compartment_id
        ):
            if instance.lifecycle_state == "TERMINATED":
                continue
            instance_by_id[instance.id] = instance
            data["instances"].append({
                "compartment": compartment_name,
                "instance": instance.display_name,
                "instance_ocid": instance.id,
                "state": instance.lifecycle_state,
                "availability_domain": instance.availability_domain,
                "shape": instance.shape,
            })
            for attachment in get_all(
                f"boot attachments/{instance.display_name}",
                compute.list_boot_volume_attachments,
                availability_domain=instance.availability_domain,
                compartment_id=compartment_id,
                instance_id=instance.id,
            ):
                boot_to_instance[attachment.boot_volume_id] = instance.id
            for attachment in get_all(
                f"block attachments/{instance.display_name}",
                compute.list_volume_attachments,
                compartment_id=compartment_id,
                instance_id=instance.id,
            ):
                volume_to_instance[attachment.volume_id] = instance.id

    # Boot/block volumes and source-side replication to Vinhedo (or chosen DR).
    ads = get_all("availability domains", identity.list_availability_domains, tenancy_id)
    for compartment_id, compartment_name in scope:
        for ad in ads:
            for boot in get_all(
                f"boot volumes/{compartment_name}/{ad.name}",
                block.list_boot_volumes,
                availability_domain=ad.name,
                compartment_id=compartment_id,
            ):
                instance = instance_by_id.get(boot_to_instance.get(boot.id))
                replicas = getattr(boot, "boot_volume_replicas", None) or []
                data["boot_volumes"].append({
                    "compartment": compartment_name,
                    "instance": getattr(instance, "display_name", ""),
                    "boot_volume": boot.display_name,
                    "boot_volume_ocid": boot.id,
                    "size_gb": boot.size_in_gbs,
                    "state": boot.lifecycle_state,
                    "direct_policy": policy_by_asset.get(boot.id, ""),
                    "replica_count": len(replicas),
                })
                for replica in replicas:
                    rid = getattr(replica, "boot_volume_replica_id", "") or ""
                    parts = rid.split(".")
                    target_region = parts[3] if len(parts) > 3 else ""
                    data["boot_replicas"].append({
                        "compartment": compartment_name,
                        "instance": getattr(instance, "display_name", ""),
                        "source_region": args.source_region,
                        "source_boot_volume": boot.display_name,
                        "source_boot_volume_ocid": boot.id,
                        "target_region": target_region,
                        "target_ad": getattr(replica, "availability_domain", ""),
                        "replica_name": getattr(replica, "display_name", ""),
                        "replica_ocid": rid,
                    })

        for volume in get_all(
            f"block volumes/{compartment_name}", block.list_volumes, compartment_id=compartment_id
        ):
            instance = instance_by_id.get(volume_to_instance.get(volume.id))
            data["block_volumes"].append({
                "compartment": compartment_name,
                "instance": getattr(instance, "display_name", ""),
                "block_volume": volume.display_name,
                "block_volume_ocid": volume.id,
                "size_gb": volume.size_in_gbs,
                "state": volume.lifecycle_state,
                "direct_policy": policy_by_asset.get(volume.id, ""),
            })

        for group in get_all(
            f"volume groups/{compartment_name}", block.list_volume_groups, compartment_id=compartment_id
        ):
            replicas = getattr(group, "volume_group_replicas", None) or []
            data["volume_groups"].append({
                "compartment": compartment_name,
                "volume_group": group.display_name,
                "volume_group_ocid": group.id,
                "state": group.lifecycle_state,
                "backup_policy": policy_by_asset.get(group.id, ""),
                "volume_count": len(group.volume_ids or []),
                "volume_ocids": ",".join(group.volume_ids or []),
                "replica_count": len(replicas),
            })
            for replica in replicas:
                rid = getattr(replica, "volume_group_replica_id", "") or ""
                parts = rid.split(".")
                target_region = parts[3] if len(parts) > 3 else ""
                data["vg_replicas"].append({
                    "compartment": compartment_name,
                    "source_region": args.source_region,
                    "source_volume_group": group.display_name,
                    "source_volume_group_ocid": group.id,
                    "target_region": target_region,
                    "target_ad": getattr(replica, "availability_domain", ""),
                    "replica_name": getattr(replica, "display_name", ""),
                    "replica_ocid": rid,
                })

    # Volume Group backups are the main backup mechanism identified in this environment.
    for region in (args.source_region, args.dr_region):
        regional_block = make_client(oci.core.BlockstorageClient, config, region)
        for compartment_id, compartment_name in scope:
            for backup in get_all(
                f"VG backups/{region}/{compartment_name}",
                regional_block.list_volume_group_backups,
                compartment_id=compartment_id,
            ):
                data["vg_backups"].append({
                    "region": region,
                    "compartment": compartment_name,
                    "backup_name": backup.display_name,
                    "backup_ocid": backup.id,
                    "volume_group_ocid": getattr(backup, "volume_group_id", ""),
                    "backup_type": getattr(backup, "type", ""),
                    "source_type": getattr(backup, "source_type", ""),
                    "state": backup.lifecycle_state,
                    "size_gb": getattr(backup, "size_in_gbs", ""),
                    "created": iso(getattr(backup, "time_created", None)),
                    "expiration": iso(getattr(backup, "expiration_time", None)),
                })

    # Review-only findings. They do not automatically mean non-compliance.
    dr_boots = {
        r["source_boot_volume_ocid"] for r in data["boot_replicas"]
        if r["target_region"] == args.dr_region
    }
    dr_groups = {
        r["source_volume_group_ocid"] for r in data["vg_replicas"]
        if r["target_region"] == args.dr_region
    }
    for row in data["boot_volumes"]:
        if not row["direct_policy"]:
            data["gaps"].append({
                "severity": "REVIEW",
                "resource_type": "BOOT_VOLUME",
                "resource": row["boot_volume"],
                "instance": row["instance"],
                "finding": "No direct policy discovered; validate whether protection is provided through a Volume Group.",
            })
        if row["boot_volume_ocid"] not in dr_boots:
            data["gaps"].append({
                "severity": "INFO",
                "resource_type": "BOOT_VOLUME",
                "resource": row["boot_volume"],
                "instance": row["instance"],
                "finding": f"No direct Boot Volume replica discovered to {args.dr_region}; validate Volume Group replication.",
            })
    for row in data["volume_groups"]:
        if not row["backup_policy"]:
            data["gaps"].append({
                "severity": "REVIEW",
                "resource_type": "VOLUME_GROUP",
                "resource": row["volume_group"],
                "instance": "",
                "finding": "No Backup Policy assignment discovered for this Volume Group.",
            })
        if row["volume_group_ocid"] not in dr_groups:
            data["gaps"].append({
                "severity": "INFO",
                "resource_type": "VOLUME_GROUP",
                "resource": row["volume_group"],
                "instance": "",
                "finding": f"No Volume Group replica discovered to {args.dr_region}.",
            })

    summary = [
        {"metric": "Tenancy", "value": tenancy.name},
        {"metric": "Source region", "value": args.source_region},
        {"metric": "DR region", "value": args.dr_region},
        {"metric": "Accessible compartments incl. root", "value": len(scope)},
        {"metric": "Instances", "value": len(data["instances"])},
        {"metric": "Backup policies", "value": len(data["policies"])},
        {"metric": "Policy schedules", "value": len(data["schedules"])},
        {"metric": "Policy assignments", "value": len(data["assignments"])},
        {"metric": "Volume Groups", "value": len(data["volume_groups"])},
        {"metric": "Volume Group backups", "value": len(data["vg_backups"])},
        {"metric": f"Boot replicas to {args.dr_region}", "value": len(dr_boots)},
        {"metric": f"VG replicas to {args.dr_region}", "value": len(dr_groups)},
        {"metric": "Items for review", "value": len(data["gaps"])},
        {"metric": "API warnings", "value": len(ERRORS)},
        {"metric": "Generated UTC", "value": datetime.now(timezone.utc).isoformat()},
        {"metric": "Execution mode", "value": "READ-ONLY"},
    ]

    sheets = [
        ("00_Resumo", summary),
        ("01_Instancias", data["instances"]),
        ("02_Boot_Volumes", data["boot_volumes"]),
        ("03_Block_Volumes", data["block_volumes"]),
        ("04_Volume_Groups", data["volume_groups"]),
        ("05_Backup_Policies", data["policies"]),
        ("06_Policy_Schedules", data["schedules"]),
        ("07_Assignments", data["assignments"]),
        ("08_VG_Backups", data["vg_backups"]),
        ("09_DR_Boot_Replicas", data["boot_replicas"]),
        ("10_DR_VG_Replicas", data["vg_replicas"]),
        ("11_Gaps", data["gaps"]),
        ("12_Errors", ERRORS),
    ]

    stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    base = f"{args.prefix}_{stamp}"
    for name, rows in sheets:
        write_csv(output_dir / f"{base}_{name}.csv", rows)
    xlsx = output_dir / f"{base}.xlsx"
    xlsx_ok = write_xlsx(xlsx, sheets)

    print("\n=== OCI BACKUP & DR AUDIT COMPLETE ===")
    for name, rows in sheets:
        print(f"{name:24}: {len(rows)}")
    print(f"Output directory        : {output_dir}")
    if xlsx_ok:
        print(f"Excel report            : {xlsx}")
    print("Execution mode          : READ-ONLY")


if __name__ == "__main__":
    main()

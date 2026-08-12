#!/usr/bin/env python3
"""
OCI Audit Environment Inventory

Read-only inventory focused on audit evidence for environment segregation.
Collects, across subscribed OCI regions:
  - Compartments
  - Compute instances with IPs, VCN/Subnet, NSGs, tags and sizing
  - VCNs and subnets from the Network compartment

Outputs:
  - oci_audit_environment_inventory_<timestamp>.xlsx
  - oci_audit_instances_<timestamp>.csv
  - oci_audit_compartments_<timestamp>.csv
  - oci_audit_network_<timestamp>.csv

Designed to run from OCI Cloud Shell or a workstation with ~/.oci/config.
No OCI resources are created, modified, stopped or deleted.
"""

import csv
import json
import logging
import os
import sys
from datetime import datetime
from pathlib import Path

import oci

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    from openpyxl.utils import get_column_letter
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False


NETWORK_COMPARTMENT_NAME = "Network"
# Empty = collect all subscribed regions. This automatically includes Sao Paulo
# and Vinhedo when the tenancy is subscribed to them.
TARGET_REGIONS = []
OUTPUT_DIR = Path("oci_audit_inventory_output")

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s - %(levelname)s - %(message)s",
    handlers=[logging.StreamHandler(sys.stdout)],
)


def load_auth():
    """Load OCI SDK authentication, preferring Cloud Shell security token."""
    try:
        signer = oci.auth.signers.get_resource_principals_signer()
        region = os.environ.get("OCI_REGION")
        if region:
            logging.info("Using Resource Principal authentication.")
            return {"region": region}, signer
    except Exception:
        pass

    try:
        config = oci.config.from_file()
        logging.info("Using OCI config from ~/.oci/config.")
        return config, None
    except Exception as exc:
        logging.error("Unable to load OCI authentication: %s", exc)
        logging.error("Run from OCI Cloud Shell or configure ~/.oci/config.")
        sys.exit(1)


def client(client_class, config, signer=None, region=None):
    cfg = dict(config)
    if region:
        cfg["region"] = region
    if signer:
        return client_class(cfg, signer=signer)
    return client_class(cfg)


def json_text(value):
    return json.dumps(value or {}, ensure_ascii=False, sort_keys=True)


def safe(value, default=""):
    return default if value is None else value


def get_tenancy_id(config, signer):
    if config.get("tenancy"):
        return config["tenancy"]
    tenancy_id = getattr(signer, "tenancy_id", None)
    if tenancy_id:
        return tenancy_id
    raise RuntimeError("Unable to determine tenancy OCID from authentication context.")


def list_compartments(identity_client, tenancy_id):
    """Return root + active compartments and build full compartment paths."""
    raw = oci.pagination.list_call_get_all_results(
        identity_client.list_compartments,
        tenancy_id,
        compartment_id_in_subtree=True,
        access_level="ANY",
    ).data

    root_name = identity_client.get_tenancy(tenancy_id).data.name
    by_id = {c.id: c for c in raw}

    def path_for(compartment):
        parts = [compartment.name]
        parent_id = compartment.compartment_id
        visited = set()
        while parent_id and parent_id != tenancy_id and parent_id not in visited:
            visited.add(parent_id)
            parent = by_id.get(parent_id)
            if not parent:
                break
            parts.append(parent.name)
            parent_id = parent.compartment_id
        return root_name + " (root)/" + "/".join(reversed(parts))

    rows = [{
        "Compartment Name": root_name,
        "Compartment Path": root_name + " (root)",
        "Parent Compartment": "",
        "Lifecycle State": "ACTIVE",
        "Description": "Tenancy root",
        "Created": "",
        "OCID": tenancy_id,
    }]

    active = []
    for c in raw:
        if c.lifecycle_state != "ACTIVE":
            continue
        active.append(c)
        parent = by_id.get(c.compartment_id)
        rows.append({
            "Compartment Name": c.name,
            "Compartment Path": path_for(c),
            "Parent Compartment": parent.name if parent else (root_name if c.compartment_id == tenancy_id else ""),
            "Lifecycle State": c.lifecycle_state,
            "Description": safe(c.description),
            "Created": c.time_created.isoformat() if c.time_created else "",
            "OCID": c.id,
        })

    # Lightweight object representing tenancy root for regional queries.
    root = oci.identity.models.Compartment(id=tenancy_id, name=root_name, compartment_id=None)
    active.append(root)
    return active, rows, {row["OCID"]: row["Compartment Path"] for row in rows}


def get_regions(identity_client, tenancy_id):
    regions = [r.region_name for r in identity_client.list_region_subscriptions(tenancy_id).data]
    if TARGET_REGIONS:
        wanted = set(TARGET_REGIONS)
        regions = [r for r in regions if r in wanted]
    return sorted(regions)


def get_vnic_context(compute_client, network_client, compartment_id, instance_id):
    private_ips, public_ips, subnet_names, vcn_names, nsg_names = [], [], [], [], []
    seen_subnets, seen_vcns, seen_nsgs = {}, {}, {}

    attachments = oci.pagination.list_call_get_all_results(
        compute_client.list_vnic_attachments,
        compartment_id=compartment_id,
        instance_id=instance_id,
    ).data

    for attachment in attachments:
        vnic = network_client.get_vnic(attachment.vnic_id).data
        if vnic.private_ip:
            private_ips.append(vnic.private_ip)
        if vnic.public_ip:
            public_ips.append(vnic.public_ip)

        if vnic.subnet_id:
            if vnic.subnet_id not in seen_subnets:
                subnet = network_client.get_subnet(vnic.subnet_id).data
                seen_subnets[vnic.subnet_id] = subnet
            subnet = seen_subnets[vnic.subnet_id]
            subnet_names.append(subnet.display_name)

            if subnet.vcn_id:
                if subnet.vcn_id not in seen_vcns:
                    vcn = network_client.get_vcn(subnet.vcn_id).data
                    seen_vcns[subnet.vcn_id] = vcn
                vcn_names.append(seen_vcns[subnet.vcn_id].display_name)

        for nsg_id in vnic.nsg_ids or []:
            if nsg_id not in seen_nsgs:
                try:
                    nsg = network_client.get_network_security_group(nsg_id).data
                    seen_nsgs[nsg_id] = nsg.display_name
                except Exception:
                    seen_nsgs[nsg_id] = nsg_id
            nsg_names.append(seen_nsgs[nsg_id])

    return {
        "Private IP": ", ".join(sorted(set(private_ips))),
        "Public IP": ", ".join(sorted(set(public_ips))),
        "VCN": ", ".join(sorted(set(vcn_names))),
        "Subnet": ", ".join(sorted(set(subnet_names))),
        "NSGs": ", ".join(sorted(set(nsg_names))),
    }


def collect_instances(regions, compartments, compartment_paths, config, signer):
    rows = []
    for region in regions:
        logging.info("Collecting Compute inventory in region %s", region)
        compute = client(oci.core.ComputeClient, config, signer, region)
        network = client(oci.core.VirtualNetworkClient, config, signer, region)

        for compartment in compartments:
            try:
                instances = oci.pagination.list_call_get_all_results(
                    compute.list_instances,
                    compartment.id,
                ).data
            except oci.exceptions.ServiceError as exc:
                logging.warning("Cannot list instances in %s/%s: %s", region, compartment.name, exc.message)
                continue

            for instance in instances:
                if instance.lifecycle_state == "TERMINATED":
                    continue

                logging.info("  %s / %s / %s", region, compartment.name, instance.display_name)
                try:
                    net = get_vnic_context(compute, network, compartment.id, instance.id)
                except Exception as exc:
                    logging.warning("    Network details unavailable for %s: %s", instance.display_name, exc)
                    net = {"Private IP": "", "Public IP": "", "VCN": "", "Subnet": "", "NSGs": ""}

                os_name, os_version = "", ""
                try:
                    image_id = getattr(instance, "image_id", None)
                    if image_id:
                        image = compute.get_image(image_id).data
                        os_name = safe(image.operating_system)
                        os_version = safe(image.operating_system_version)
                except Exception:
                    pass

                shape_config = getattr(instance, "shape_config", None)
                rows.append({
                    "Region": region,
                    "Compartment": compartment.name,
                    "Compartment Path": compartment_paths.get(compartment.id, compartment.name),
                    "Instance Name": instance.display_name,
                    "State": instance.lifecycle_state,
                    "Shape": instance.shape,
                    "OCPUs": safe(getattr(shape_config, "ocpus", None)),
                    "Memory (GB)": safe(getattr(shape_config, "memory_in_gbs", None)),
                    "Availability Domain": safe(instance.availability_domain),
                    "Fault Domain": safe(instance.fault_domain),
                    "VCN": net["VCN"],
                    "Subnet": net["Subnet"],
                    "Private IP": net["Private IP"],
                    "Public IP": net["Public IP"],
                    "NSGs": net["NSGs"],
                    "Operating System": os_name,
                    "OS Version": os_version,
                    "Created": instance.time_created.isoformat() if instance.time_created else "",
                    "Freeform Tags": json_text(instance.freeform_tags),
                    "Defined Tags": json_text(instance.defined_tags),
                    "OCID": instance.id,
                })
    return rows


def find_network_compartments(compartments, compartment_paths):
    matches = []
    for c in compartments:
        if c.name.lower() == NETWORK_COMPARTMENT_NAME.lower():
            matches.append(c)
    if not matches:
        logging.warning("No active compartment named '%s' was found.", NETWORK_COMPARTMENT_NAME)
    else:
        for c in matches:
            logging.info("Network compartment selected: %s", compartment_paths.get(c.id, c.name))
    return matches


def collect_network(regions, network_compartments, compartment_paths, config, signer):
    rows = []
    for region in regions:
        network = client(oci.core.VirtualNetworkClient, config, signer, region)
        logging.info("Collecting VCN/Subnet inventory in region %s", region)

        for compartment in network_compartments:
            try:
                vcns = oci.pagination.list_call_get_all_results(
                    network.list_vcns,
                    compartment.id,
                ).data
            except oci.exceptions.ServiceError as exc:
                logging.warning("Cannot list VCNs in %s/%s: %s", region, compartment.name, exc.message)
                continue

            for vcn in vcns:
                if vcn.lifecycle_state == "TERMINATED":
                    continue

                cidrs = getattr(vcn, "cidr_blocks", None) or ([vcn.cidr_block] if vcn.cidr_block else [])
                rows.append({
                    "Region": region,
                    "Compartment": compartment.name,
                    "Compartment Path": compartment_paths.get(compartment.id, compartment.name),
                    "Resource Type": "VCN",
                    "VCN": vcn.display_name,
                    "Resource Name": vcn.display_name,
                    "CIDR": ", ".join(cidrs),
                    "Public/Private": "N/A",
                    "DNS Label": safe(vcn.dns_label),
                    "Lifecycle State": vcn.lifecycle_state,
                    "Freeform Tags": json_text(vcn.freeform_tags),
                    "Defined Tags": json_text(vcn.defined_tags),
                    "OCID": vcn.id,
                })

                try:
                    subnets = oci.pagination.list_call_get_all_results(
                        network.list_subnets,
                        compartment_id=compartment.id,
                        vcn_id=vcn.id,
                    ).data
                except oci.exceptions.ServiceError as exc:
                    logging.warning("Cannot list subnets for VCN %s: %s", vcn.display_name, exc.message)
                    continue

                for subnet in subnets:
                    if subnet.lifecycle_state == "TERMINATED":
                        continue
                    rows.append({
                        "Region": region,
                        "Compartment": compartment.name,
                        "Compartment Path": compartment_paths.get(compartment.id, compartment.name),
                        "Resource Type": "Subnet",
                        "VCN": vcn.display_name,
                        "Resource Name": subnet.display_name,
                        "CIDR": subnet.cidr_block,
                        "Public/Private": "Private" if subnet.prohibit_public_ip_on_vnic else "Public-capable",
                        "DNS Label": safe(subnet.dns_label),
                        "Lifecycle State": subnet.lifecycle_state,
                        "Freeform Tags": json_text(subnet.freeform_tags),
                        "Defined Tags": json_text(subnet.defined_tags),
                        "OCID": subnet.id,
                    })
    return rows


def write_csv(path, rows):
    if not rows:
        return
    with path.open("w", newline="", encoding="utf-8-sig") as handle:
        writer = csv.DictWriter(handle, fieldnames=list(rows[0].keys()))
        writer.writeheader()
        writer.writerows(rows)


def autosize_and_style(ws):
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    for col_idx, column in enumerate(ws.columns, 1):
        max_len = 0
        for cell in column:
            value = "" if cell.value is None else str(cell.value)
            max_len = max(max_len, min(len(value), 60))
        ws.column_dimensions[get_column_letter(col_idx)].width = max(12, max_len + 2)


def add_sheet(workbook, title, rows):
    ws = workbook.create_sheet(title=title)
    if not rows:
        ws.append(["No data collected"])
        return
    headers = list(rows[0].keys())
    ws.append(headers)
    for row in rows:
        ws.append([row.get(h, "") for h in headers])
    autosize_and_style(ws)


def write_xlsx(path, compartment_rows, instance_rows, network_rows, regions):
    if not OPENPYXL_AVAILABLE:
        logging.warning("openpyxl is not installed; XLSX output will be skipped.")
        logging.warning("Install with: pip3 install --user openpyxl")
        return

    wb = Workbook()
    summary = wb.active
    summary.title = "Summary"
    summary.append(["OCI Audit Environment Inventory", ""])
    summary.append(["Generated", datetime.now().astimezone().isoformat(timespec="seconds")])
    summary.append(["Regions collected", ", ".join(regions)])
    summary.append(["Active compartments", len(compartment_rows)])
    summary.append(["Compute instances", len(instance_rows)])
    summary.append(["Network resources (VCN/Subnet)", len(network_rows)])
    summary.append(["Network compartment filter", NETWORK_COMPARTMENT_NAME])
    summary.append(["Scope", "Read-only inventory for audit evidence of environment segregation"])
    summary.column_dimensions["A"].width = 38
    summary.column_dimensions["B"].width = 90
    summary["A1"].font = Font(bold=True, size=14)

    add_sheet(wb, "Compartments", compartment_rows)
    add_sheet(wb, "Instances", instance_rows)
    add_sheet(wb, "Network", network_rows)
    wb.save(path)


def main():
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")

    config, signer = load_auth()
    tenancy_id = get_tenancy_id(config, signer)
    identity = client(oci.identity.IdentityClient, config, signer, config.get("region"))

    logging.info("Collecting compartment hierarchy...")
    compartments, compartment_rows, compartment_paths = list_compartments(identity, tenancy_id)
    regions = get_regions(identity, tenancy_id)
    logging.info("Subscribed regions selected: %s", ", ".join(regions))

    network_compartments = find_network_compartments(compartments, compartment_paths)
    instance_rows = collect_instances(regions, compartments, compartment_paths, config, signer)
    network_rows = collect_network(regions, network_compartments, compartment_paths, config, signer)

    compartments_csv = OUTPUT_DIR / f"oci_audit_compartments_{timestamp}.csv"
    instances_csv = OUTPUT_DIR / f"oci_audit_instances_{timestamp}.csv"
    network_csv = OUTPUT_DIR / f"oci_audit_network_{timestamp}.csv"
    xlsx = OUTPUT_DIR / f"oci_audit_environment_inventory_{timestamp}.xlsx"

    write_csv(compartments_csv, compartment_rows)
    write_csv(instances_csv, instance_rows)
    write_csv(network_csv, network_rows)
    write_xlsx(xlsx, compartment_rows, instance_rows, network_rows, regions)

    logging.info("Inventory complete.")
    logging.info("Compartments: %s", compartments_csv)
    logging.info("Instances:    %s", instances_csv)
    logging.info("Network:      %s", network_csv)
    if OPENPYXL_AVAILABLE:
        logging.info("Excel:        %s", xlsx)


if __name__ == "__main__":
    main()
